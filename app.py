from flask import Flask, Response, jsonify, request, send_file
from flask_cors import CORS
import cv2
import json
import time
import csv
import io
import os
import sqlite3
from datetime import datetime, date
import threading
import numpy as np

try:
    from ultralytics import YOLO
    YOLO_AVAILABLE = True
    print("✅ YOLO loaded successfully")
except ImportError:
    YOLO_AVAILABLE = False
    print("⚠️  YOLO not installed. Running in SIMULATION mode.")
    print("   To install: pip install ultralytics")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not installed. Excel export disabled.")
    print("   To install: pip install openpyxl")

app = Flask(__name__)
CORS(app) 
DB_FILE = "counts.db"

def init_db():
    """Create the database tables if they don't exist."""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS daily_counts (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            date       TEXT NOT NULL,
            hour       INTEGER NOT NULL,
            males      INTEGER DEFAULT 0,
            females    INTEGER DEFAULT 0,
            cars       INTEGER DEFAULT 0,
            motorcycles INTEGER DEFAULT 0,
            trucks     INTEGER DEFAULT 0,
            objects    INTEGER DEFAULT 0,
            UNIQUE(date, hour)
        )
    """)
    conn.commit()
    conn.close()
    print("✅ Database ready")

init_db()

state = {
    "males":       0,
    "females":     0,
    "cars":        0,
    "motorcycles": 0,
    "trucks":      0,
    "objects":     0,
    "camera_on":   False,
    "camera_url":  "",
}

LINE_POSITION = 0.5

PERSON_CLASS = 0       
CAR_CLASS    = 2      
MOTO_CLASS   = 3       
TRUCK_CLASS  = 7      

OBJECT_CLASSES = {
    41: "cup",
    56: "chair",
    57: "couch",
    60: "dining table",
    63: "laptop",
    67: "cell phone",
    73: "book",
}

tracked_objects = {}   

def check_line_crossing(obj_id, current_y, frame_height):
    
    line_y = frame_height * LINE_POSITION

    if obj_id not in tracked_objects:
        tracked_objects[obj_id] = {"prev_y": current_y, "counted": False}
        return False

    prev_y = tracked_objects[obj_id]["prev_y"]
    already_counted = tracked_objects[obj_id]["counted"]

    crossed = (prev_y < line_y and current_y >= line_y) or \
              (prev_y > line_y and current_y <= line_y)

    tracked_objects[obj_id]["prev_y"] = current_y

    if crossed and not already_counted:
        tracked_objects[obj_id]["counted"] = True
        return True

    return False

import random

def estimate_gender():
    return "male" if random.random() > 0.5 else "female"

def save_to_db():

    today = date.today().isoformat()
    hour  = datetime.now().hour

    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO daily_counts (date, hour, males, females, cars, motorcycles, trucks, objects)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(date, hour) DO UPDATE SET
            males       = excluded.males,
            females     = excluded.females,
            cars        = excluded.cars,
            motorcycles = excluded.motorcycles,
            trucks      = excluded.trucks,
            objects     = excluded.objects
    """, (today, hour,
          state["males"], state["females"],
          state["cars"],  state["motorcycles"],
          state["trucks"], state["objects"]))

    conn.commit()
    conn.close()

camera_thread = None
camera_active = False
latest_frame  = None   

def camera_worker(url):
    """
    Background thread:
    1. Opens the camera (IP or webcam)
    2. Runs YOLO object detection on each frame
    3. Applies line crossing logic
    4. Updates global state counts
    5. Draws boxes and the counting line on the frame
    """
    global latest_frame, camera_active, state

    print(f"📷 Opening camera: {url}")
    cap = cv2.VideoCapture(url)

    if not cap.isOpened():
        print(f"❌ Cannot open camera: {url}")
        state["camera_on"] = False
        return

    model = YOLO("yolov8n.pt") if YOLO_AVAILABLE else None
    if model:
        print("🤖 YOLO model loaded")

    frame_id = 0

    while camera_active:
        ret, frame = cap.read()
        if not ret:
            print("⚠️  Lost camera connection. Retrying...")
            time.sleep(2)
            cap.open(url)
            continue

        h, w = frame.shape[:2]
        line_y = int(h * LINE_POSITION)

        cv2.line(frame, (0, line_y), (w, line_y), (0, 255, 255), 2)
        cv2.putText(frame, "Counting Line", (10, line_y - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2)

        if model:
            results = model.track(frame, persist=True, verbose=False)

            if results and results[0].boxes is not None:
                boxes = results[0].boxes

                for box in boxes:
                    x1, y1, x2, y2 = map(int, box.xyxy[0])
                    cls_id   = int(box.cls[0])       # Class ID
                    conf     = float(box.conf[0])    # Confidence (0-1)
                    obj_id   = int(box.id[0]) if box.id is not None else frame_id

                    if conf < 0.4:   
                        continue

                    center_y = (y1 + y2) // 2      

                    crossed = check_line_crossing(obj_id, center_y, h)

                    if crossed:
                        
                        if cls_id == PERSON_CLASS:
                            gender = estimate_gender()
                            if gender == "male":
                                state["males"] += 1
                            else:
                                state["females"] += 1
                        elif cls_id == CAR_CLASS:
                            state["cars"] += 1
                        elif cls_id == MOTO_CLASS:
                            state["motorcycles"] += 1
                        elif cls_id == TRUCK_CLASS:
                            state["trucks"] += 1
                        elif cls_id in OBJECT_CLASSES:
                            state["objects"] += 1

                        save_to_db()  

                    color = (0, 255, 0)   
                    label = "unknown"

                    if cls_id == PERSON_CLASS:
                        color = (255, 100, 0)
                        label = "Person"
                    elif cls_id == CAR_CLASS:
                        color = (0, 200, 255)
                        label = "Car"
                    elif cls_id == MOTO_CLASS:
                        color = (0, 100, 255)
                        label = "Motorcycle"
                    elif cls_id == TRUCK_CLASS:
                        color = (255, 0, 100)
                        label = "Truck"
                    elif cls_id in OBJECT_CLASSES:
                        color = (200, 200, 0)
                        label = OBJECT_CLASSES[cls_id]

                    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                    cv2.putText(frame, f"{label} {conf:.1f}",
                                (x1, y1 - 8),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

        overlay_text = [
            f"People: {state['males'] + state['females']} (M:{state['males']} F:{state['females']})",
            f"Cars: {state['cars']}  Motos: {state['motorcycles']}  Trucks: {state['trucks']}",
        ]
        for i, txt in enumerate(overlay_text):
            cv2.putText(frame, txt, (10, 30 + i * 25),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

        _, jpeg = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
        latest_frame = jpeg.tobytes()

        frame_id += 1
        time.sleep(0.03)  

    cap.release()
    print("📷 Camera closed")

def simulation_worker():
    """Generates fake detection data every few seconds for testing."""
    global camera_active, state, latest_frame

    print("🎭 Simulation mode started")

    frame = np.zeros((480, 640, 3), dtype=np.uint8)
    h, w  = frame.shape[:2]

    while camera_active:
        time.sleep(random.uniform(2, 4))
        if not camera_active:
            break

        choice = random.choice(["male", "female", "car", "motorcycle", "truck", "object"])
        if choice == "male":
            state["males"] += 1
        elif choice == "female":
            state["females"] += 1
        elif choice == "car":
            state["cars"] += 1
        elif choice == "motorcycle":
            state["motorcycles"] += 1
        elif choice == "truck":
            state["trucks"] += 1
        else:
            state["objects"] += 1

        save_to_db()

        sim = frame.copy()
        line_y = int(h * LINE_POSITION)
        cv2.line(sim, (0, line_y), (w, line_y), (0, 255, 255), 2)
        cv2.putText(sim, "SIMULATION MODE - No Real Camera",
                    (80, h // 2), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 200, 255), 2)
        cv2.putText(sim, f"People: {state['males']+state['females']}  Cars: {state['cars']}",
                    (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

        _, jpeg = cv2.imencode(".jpg", sim)
        latest_frame = jpeg.tobytes()

@app.route("/api/start-camera", methods=["POST"])
def start_camera():
    global camera_thread, camera_active, state

    data = request.json
    url  = data.get("url", "")

    if camera_active:
        return jsonify({"error": "Camera already running"}), 400

    state.update({
        "males": 0, "females": 0, "cars": 0,
        "motorcycles": 0, "trucks": 0, "objects": 0,
        "camera_on": True, "camera_url": url
    })

    camera_active = True

    if url and not url.startswith("SIM"):
        target = camera_worker
        args   = (url,)
    else:
        target = simulation_worker
        args   = ()

    camera_thread = threading.Thread(target=target, args=args, daemon=True)
    camera_thread.start()

    return jsonify({"status": "Camera started", "mode": "real" if url else "simulation"})


@app.route("/api/stop-camera", methods=["POST"])
def stop_camera():
    """Stop the camera feed."""
    global camera_active
    camera_active = False
    state["camera_on"] = False
    return jsonify({"status": "Camera stopped"})


@app.route("/api/counts")
def get_counts():
    """Return live counts to the frontend dashboard."""
    people = state["males"] + state["females"]
    vehicles = state["cars"] + state["motorcycles"] + state["trucks"]
    return jsonify({
        "males":        state["males"],
        "females":      state["females"],
        "cars":         state["cars"],
        "motorcycles":  state["motorcycles"],
        "trucks":       state["trucks"],
        "objects":      state["objects"],
        "total_people":      people,
        "total_vehicles":    vehicles,
        "camera_on":    state["camera_on"],
    })


@app.route("/api/reset", methods=["POST"])
def reset_counts():
    """Reset all counts to zero."""
    state.update({"males": 0, "females": 0, "cars": 0,
                  "motorcycles": 0, "trucks": 0, "objects": 0})
    return jsonify({"status": "Counts reset"})


@app.route("/video-feed")
def video_feed():

    def generate():
        while True:
            if latest_frame:
                yield (b"--frame\r\n"
                       b"Content-Type: image/jpeg\r\n\r\n" +
                       latest_frame + b"\r\n")
            time.sleep(0.04)

    return Response(generate(),
                    mimetype="multipart/x-mixed-replace; boundary=frame")


@app.route("/api/daily-report")
def daily_report():
    today = date.today().isoformat()
    conn  = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT hour, males, females, cars, motorcycles, trucks, objects
        FROM daily_counts
        WHERE date = ?
        ORDER BY hour
    """, (today,))
    rows = cursor.fetchall()
    conn.close()

    data = []
    for row in rows:
        hour, males, females, cars, motos, trucks, objects = row
        data.append({
            "hour":         hour,
            "label":        f"{hour:02d}:00",
            "males":        males,
            "females":      females,
            "cars":         cars,
            "motorcycles":  motos,
            "trucks":       trucks,
            "objects":      objects,
            "total_people": males + females,
        })

    return jsonify({"date": today, "hourly": data})


@app.route("/api/export-csv")
def export_csv():
    """Export today's data as a CSV file."""
    today = date.today().isoformat()
    conn  = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT date, hour, males, females, cars, motorcycles, trucks, objects
        FROM daily_counts WHERE date = ?
        ORDER BY hour
    """, (today,))
    rows = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["Date", "Hour", "Males", "Females",
                     "Total People", "Cars", "Motorcycles", "Trucks", "Objects"])

    for row in rows:
        date_val, hour, males, females, cars, motos, trucks, objects = row
        writer.writerow([date_val, f"{hour:02d}:00",
                         males, females, males + females,
                         cars, motos, trucks, objects])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=report_{today}.csv"}
    )


@app.route("/api/export-excel")
def export_excel():
    """Export today's data as an Excel (.xlsx) file."""
    if not EXCEL_AVAILABLE:
        return jsonify({"error": "openpyxl not installed"}), 500

    today = date.today().isoformat()
    conn  = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT date, hour, males, females, cars, motorcycles, trucks, objects
        FROM daily_counts WHERE date = ?
        ORDER BY hour
    """, (today,))
    rows = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Report"

    headers = ["Date", "Hour", "Males", "Females",
               "Total People", "Cars", "Motorcycles", "Trucks", "Objects"]
    header_fill = PatternFill("solid", fgColor="1a73e8")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        date_val, hour, males, females, cars, motos, trucks, objects = row
        ws.append([date_val, f"{hour:02d}:00",
                   males, females, males + females,
                   cars, motos, trucks, objects])

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"report_{today}.xlsx"
    )


if __name__ == "__main__":
    print("\n🚀 Starting Object Counting Backend Server...")
    print("   Open http://localhost:5000 in your browser after starting frontend\n")
    app.run(host="0.0.0.0", port=5000, debug=False)
